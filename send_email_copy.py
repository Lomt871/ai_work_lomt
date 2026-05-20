import os
import time
import smtplib
import shutil
from typing import List, Optional, Tuple
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from playwright.sync_api import sync_playwright


def log(message):
    """记录日志到文件"""
    log_path = os.path.join(os.path.dirname(__file__), 'send_email.log')
    with open(log_path, 'a', encoding='utf-8') as f:
        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")


def format_value(value, col_index):
    """格式化单元格值"""
    if value is None:
        return ""
    
    # 处理数字类型
    if isinstance(value, (int, float)):
        # 学员id、课包id、课时数、支付前未开课课包数、月初是否不可续去掉.0
        if col_index in [1, 7, 9, 10, 11]:  # 对应列索引
            if value == int(value):
                return str(int(value))
        # 续费课包金额使用千分位
        elif col_index == 6:  # 续费课包金额
            return "{:,.0f}".format(value)
    
    # 处理日期类型 - 直接输出Excel中的日期
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    
    return str(value)


def parse_email_list(raw_emails: str) -> List[str]:
    return [email.strip() for email in raw_emails.split(',') if email.strip()]


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "")


def extract_report_stats(file_path: str) -> Tuple[int, str]:
    order_count = 0
    order_month = ""

    if file_path.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        learner_col_idx: Optional[int] = None
        data_start_row = 4

        max_scan_row = min(ws.max_row, 10)
        for row_idx in range(1, max_scan_row + 1):
            row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
            normalized = [normalize_header(v) for v in row_values]
            if "学员id" in normalized:
                learner_col_idx = normalized.index("学员id") + 1
                data_start_row = row_idx + 1
                break

        if learner_col_idx is None:
            learner_col_idx = 2

        for row_idx in range(data_start_row, ws.max_row + 1):
            learner_id = ws.cell(row=row_idx, column=learner_col_idx).value
            if learner_id is not None and str(learner_id).strip() != "":
                order_count += 1

            pay_time = ws.cell(row=row_idx, column=6).value
            if not order_month and isinstance(pay_time, datetime):
                order_month = pay_time.strftime('%Y年%m月')

        return order_count, order_month

    if file_path.endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        learner_col_idx = None
        data_start_row = 3

        max_scan_row = min(ws.nrows, 10)
        for row_idx in range(max_scan_row):
            row_values = [ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)]
            normalized = [normalize_header(v) for v in row_values]
            if "学员id" in normalized:
                learner_col_idx = normalized.index("学员id")
                data_start_row = row_idx + 1
                break

        if learner_col_idx is None:
            learner_col_idx = 1

        for row_idx in range(data_start_row, ws.nrows):
            learner_id = ws.cell_value(row_idx, learner_col_idx)
            if str(learner_id).strip() != "":
                order_count += 1

            pay_time = ws.cell_value(row_idx, 5)
            if not order_month and isinstance(pay_time, datetime):
                order_month = pay_time.strftime('%Y年%m月')

        return order_count, order_month

    return order_count, order_month


def cleanup_existing_reports(download_dir: str) -> None:
    removed_files = []
    for file_name in os.listdir(download_dir):
        if (file_name.endswith('.xlsx') or file_name.endswith('.xls')) and not file_name.startswith('~$'):
            file_path = os.path.join(download_dir, file_name)
            try:
                os.remove(file_path)
                removed_files.append(file_name)
            except Exception as e:
                log(f"删除旧报表失败: {file_path}, 错误: {e}")

    if removed_files:
        log(f"已删除旧报表: {', '.join(removed_files)}")
        print(f"已删除旧报表: {', '.join(removed_files)}")


def read_excel_to_html(file_path):
    """读取Excel文件内容并转换为HTML表格"""
    html_table = ""
    try:
        # 根据文件扩展名选择读取方式
        if file_path.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        elif file_path.endswith('.xls'):
            import xlrd
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
        else:
            return "不支持的文件格式"

        # 构建HTML表格（内容居中，删除第一列空列）
        html_table = "<table border='1' cellpadding='4' cellspacing='0' style='border-collapse: collapse; font-size: 10.5pt;'>\n"
        
        if file_path.endswith('.xlsx'):
            # 读取所有行，跳过第一行（日期行）和空行
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                # 跳过第一行（日期行）
                if row_idx == 0:
                    continue
                
                # 跳过空行（所有单元格都是空的）
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                html_table += "  <tr>\n"
                # 跳过第一列（空列），从第2列开始
                for col_idx, cell in enumerate(row):
                    if col_idx == 0:  # 跳过第一列
                        continue
                    cell_value = format_value(cell, col_idx)
                    html_table += f"    <td align='center'>{cell_value}</td>\n"
                html_table += "  </tr>\n"
        else:
            # .xls格式
            for row_idx in range(ws.nrows):
                # 跳过第一行（日期行）
                if row_idx == 0:
                    continue
                
                html_table += "  <tr>\n"
                # 跳过第一列（空列）
                for col_idx in range(ws.ncols):
                    if col_idx == 0:  # 跳过第一列
                        continue
                    cell_value = format_value(ws.cell_value(row_idx, col_idx), col_idx)
                    html_table += f"    <td align='center'>{cell_value}</td>\n"
                html_table += "  </tr>\n"
        
        html_table += "</table>"
        
    except ImportError as e:
        html_table = f"缺少Excel解析库: {e}"
    except Exception as e:
        html_table = f"读取Excel失败: {e}"
    
    return html_table


def send_email_with_attachment(attachment_path):
    """发送带有附件和Excel内容的邮件"""
    # 邮件配置
    sender_email = "liuxiufang@hltn.com"
    sender_password = "Asd20220516"
    receiver_email = "ya.liu@hltn.com,sunhaotian@hltn.com,bijinling@hltn.com,wuminhao@hltn.com,chenmingxin@hltn.com,fangyu.zeng@hltn.com,zjb-gzc@hltn.com,wanghaitong@hltn.com,wangxiaoyu02@hltn.com"
    cc_email = "yuanwen.zhan@hltn.com,xiaoyan.liu@hltn.com,huangyaying@hltn.com,dongyi01@hltn.com,huxiaolin@hltn.com"

    # SMTP服务器配置 (腾讯企业邮箱)
    smtp_server = "smtp.exmail.qq.com"
    smtp_port = 465  # SSL端口

    try:
        order_count, order_month = extract_report_stats(attachment_path)

        # 创建邮件对象
        msg = MIMEMultipart()
        msg['From'] = sender_email
        to_emails = parse_email_list(receiver_email)
        cc_emails = parse_email_list(cc_email)
        all_recipients = list(dict.fromkeys(to_emails + cc_emails))

        if not all_recipients:
            raise ValueError("收件人和抄送人为空，无法发送邮件")

        msg['To'] = ", ".join(to_emails)
        msg['Cc'] = ", ".join(cc_emails)
        # 如果未读取到月份，使用前一天的日期
        if not order_month:
            prev_day = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
            order_month = prev_day.strftime('%Y年%m月')
        subject_prefix = "【无需确认】" if order_count == 0 else "【待确认】"
        msg['Subject'] = f"{subject_prefix}{order_month}思维不可续&未开课续费订单反查"

        # 邮件正文 (HTML格式)
        if order_count == 0:
            body = f"""
<html>
<head>
<style>
    body {{ font-family: '微软雅黑', 'Microsoft YaHei', sans-serif; font-size: 10.5pt; margin: 0; padding: 0; line-height: 1.4; }}
    p {{ margin: 2px 0; }}
</style>
</head>
<body>
<p>Dear all,</p>
<p style="text-indent: 2em;">本月无不可续&amp;未开课续费订单，无需确认说明。</p>
</body>
</html>
"""
        else:
            excel_html = read_excel_to_html(attachment_path)
            body = f"""
<html>
<head>
<style>
    body {{ font-family: '微软雅黑', 'Microsoft YaHei', sans-serif; font-size: 10.5pt; margin: 0; padding: 0; line-height: 1.4; }}
    p {{ margin: 2px 0; }}
    table {{ border-collapse: collapse; font-family: '微软雅黑', 'Microsoft YaHei', sans-serif; font-size: 10.5pt; margin-top: 4px; }}
    th, td {{ border: 1px solid #ccc; padding: 4px 8px; text-align: center; }}
    th {{ background-color: #f5f5f5; font-weight: bold; }}
</style>
</head>
<body>
<p>Dear all,</p>
<p style="text-indent: 2em;">基于续费产出健康度的考虑，月初不可续学员及存在未开课课包的学员均不在当月可续范围内，若存在特殊情况需要续费，应先经业务负责人评估合理后续费，其余未经报备或评估不合规的续费订单，一律不计入续费提成。</p>
<p style="text-indent: 2em;">现排查到{order_month}存在不可续&amp;未开课续费的订单共 {order_count} 单，烦请孙昊天老师逐单说明具体背景以及是否计算提成，之后依次由质检老师审批、刘亚老师审批。（为避免影响薪酬计算，辛苦在今天内回复~）</p>
{excel_html}
</body>
</html>
"""
        msg.attach(MIMEText(body, 'html', 'utf-8'))

        if order_count > 0:
            # 添加附件
            with open(attachment_path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())

            encoders.encode_base64(part)
            filename = os.path.basename(attachment_path)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename="{filename}"'
            )
            msg.attach(part)

        # 连接到SMTP服务器并发送邮件
        print(f"\n=== 正在发送邮件 ===")
        print(f"发件人: {sender_email}")
        print(f"收件人: {receiver_email}")
        print(f"抄送: {cc_email}")
        log(f"开始发送邮件: 发件人={sender_email}, 收件人={receiver_email}, 抄送={cc_email}")

        server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        server.login(sender_email, sender_password)

        failed_recipients = server.sendmail(sender_email, all_recipients, msg.as_string())

        if failed_recipients:
            failed_list = list(failed_recipients.keys())
            log(f"首次投递失败邮箱: {', '.join(failed_list)}，准备重试一次")
            print(f"首次投递失败邮箱: {', '.join(failed_list)}，准备重试一次")
            time.sleep(2)
            retry_failed = server.sendmail(sender_email, failed_list, msg.as_string())
            if retry_failed:
                server.quit()
                still_failed = ', '.join(retry_failed.keys())
                log(f"重试后仍失败邮箱: {still_failed}")
                raise RuntimeError(f"以下邮箱重试后仍投递失败: {still_failed}")
            log("失败邮箱重试成功")
            print("失败邮箱重试成功")

        server.quit()

        print("邮件发送成功！")
        log("邮件发送成功！")
        return True

    except Exception as e:
        print(f"发送邮件时出错: {e}")
        log(f"发送邮件时出错: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    download_dir = "D:\\不可续学员明细"

    # 确保下载目录存在
    if not os.path.exists(download_dir):
        os.makedirs(download_dir)

    print(f"下载目录: {os.path.abspath(download_dir)}")
    log("=== 开始执行任务 ===")
    log(f"下载目录: {os.path.abspath(download_dir)}")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="msedge")
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        # 下载事件处理
        download_completed = False
        download_file_path = None

        def handle_download(download):
            nonlocal download_completed, download_file_path
            print(f"\n=== 捕获到下载事件 ===")
            print(f"文件名: {download.suggested_filename}")
            print(f"下载URL: {download.url}")

            try:
                download_path = download.path()
                print(f"临时路径: {download_path}")

                target_path = os.path.join(download_dir, download.suggested_filename)

                # 删除旧文件（只删除同名文件）
                if os.path.exists(target_path):
                    os.remove(target_path)
                    print(f"删除旧文件: {target_path}")

                # 复制文件到目标目录
                shutil.copy(download_path, target_path)

                download_file_path = target_path
                download_completed = True

                print(f"文件已保存到: {target_path}")
                print(f"文件大小: {os.path.getsize(target_path)} 字节")
                log(f"文件下载成功: {target_path}, 大小: {os.path.getsize(target_path)} 字节")

            except Exception as e:
                print(f"处理下载时出错: {e}")
                log(f"处理下载时出错: {e}")

        page.on("download", handle_download)

        # 记录所有网络请求
        def handle_request(request):
            if 'export' in request.url.lower() or 'download' in request.url.lower():
                print(f"\n请求: {request.method} {request.url}")

        page.on("request", handle_request)

        # 记录所有网络响应
        def handle_response(response):
            if response.status == 404:
                print(f"\n404错误: {response.url}")
            elif 'export' in response.url.lower():
                print(f"\n响应: {response.status} {response.url}")
                content_type = response.headers.get('content-type', '')
                content_length = response.headers.get('content-length', '0')
                print(f"  Content-Type: {content_type}")
                print(f"  Content-Length: {content_length}")

        page.on("response", handle_response)

        try:
            # 登录
            print("\n=== 登录 ===")
            log("开始登录系统")
            page.goto("https://bi.61info.cn/smartbi")
            page.wait_for_load_state("load", timeout=60000)
            time.sleep(2)

            # 输入用户名
            username_input = page.wait_for_selector(".item-textinput", timeout=15000)
            username_input.fill("60220")
            print("用户名输入成功")

            # 输入密码
            password_input = page.wait_for_selector("//input[@type='password']", timeout=15000)
            password_input.fill("Asd@2025")
            print("密码输入成功")

            # 点击登录按钮
            login_button = page.wait_for_selector("//input[@bofid='login']", timeout=15000)
            login_button.click()
            print("登录按钮点击成功")
            log("登录成功")

            page.wait_for_load_state("load", timeout=60000)
            time.sleep(3)

            # 访问报表
            print("\n=== 访问报表 ===")
            log("开始访问报表")
            report_url = "https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019659bd59bdd9c301967629c3a912f2"
            page.goto(report_url)
            page.wait_for_load_state("load", timeout=60000)
            time.sleep(15)  # 增加等待时间

            # 下载前清理旧报表，避免使用到历史文件
            cleanup_existing_reports(download_dir)

            # 点击导出按钮
            print("\n=== 点击导出按钮 ===")
            log("开始导出报表")
            export_button = page.wait_for_selector("//input[@value='导出']", timeout=15000)
            page.evaluate("(element) => element.click()", export_button)
            print("导出按钮点击成功")
            time.sleep(5)

            # 找到Excel选项
            print("\n=== 找到Excel选项 ===")
            excel_option = page.wait_for_selector("//*[@id='EXCEL2007']", timeout=10000)
            print("Excel选项找到")

            # 获取Excel选项位置并移动鼠标悬停
            box = excel_option.bounding_box()
            if box:
                print(f"Excel选项位置: x={box['x']}, y={box['y']}, width={box['width']}, height={box['height']}")
                # 移动鼠标到Excel选项中心
                page.mouse.move(box['x'] + box['width'] / 2, box['y'] + box['height'] / 2)
                print("鼠标已移动到Excel选项上（悬停）")

            time.sleep(5)  # 等待在线导出选项出现

            # 查找在线导出选项
            print("\n=== 查找在线导出选项 ===")
            online_export = page.wait_for_selector("//*[@caption='在线导出']", timeout=15000, state="attached")
            print("在线导出选项找到")

            # 获取在线导出选项位置
            online_box = online_export.bounding_box()
            if online_box:
                print(f"在线导出选项位置: x={online_box['x']}, y={online_box['y']}, width={online_box['width']}, height={online_box['height']}")
                # 移动鼠标到在线导出选项上
                page.mouse.move(online_box['x'] + online_box['width'] / 2, online_box['y'] + online_box['height'] / 2)
                print("鼠标已移动到在线导出选项上")

            time.sleep(2)

            # 左键单击在线导出
            print("\n=== 左键单击在线导出 ===")
            if online_box:
                page.mouse.click(online_box['x'] + online_box['width'] / 2, online_box['y'] + online_box['height'] / 2)
            print("在线导出点击成功")

            # 同步等待下载事件
            print("\n=== 等待下载 ===")
            download_file_path = None
            
            try:
                # 使用Playwright的wait_for_event同步等待下载事件
                download = page.wait_for_event("download", timeout=300000)  # 5分钟超时
                print(f"\n=== 捕获到下载事件 ===")
                print(f"文件名: {download.suggested_filename}")
                print(f"下载URL: {download.url}")
                
                # 等待下载完成并获取下载路径
                download_path = download.path()
                print(f"临时路径: {download_path}")
                
                # 复制到目标目录
                target_path = os.path.join(download_dir, download.suggested_filename)
                
                # 删除旧文件
                if os.path.exists(target_path):
                    os.remove(target_path)
                    print(f"删除旧文件: {target_path}")
                
                shutil.copy(download_path, target_path)
                download_file_path = target_path
                
                print(f"文件已保存到: {target_path}")
                print(f"文件大小: {os.path.getsize(target_path)} 字节")
                log(f"文件下载成功: {target_path}, 大小: {os.path.getsize(target_path)} 字节")
                
            except Exception as e:
                print(f"等待下载超时或出错: {e}")
            
            # 发送邮件
            if download_file_path and os.path.exists(download_file_path):
                print(f"\n=== 准备发送邮件 ===")
                log(f"下载完成，开始编辑并发送邮件: {download_file_path}")
                send_email_with_attachment(download_file_path)
            else:
                # 尝试从目录查找文件
                print("\n=== 尝试从目录查找文件 ===")
                excel_files = [f for f in os.listdir(download_dir)
                              if (f.endswith('.xlsx') or f.endswith('.xls')) and not f.startswith('~$')]
                if excel_files:
                    latest_file = max(excel_files, key=lambda x: os.path.getmtime(os.path.join(download_dir, x)))
                    download_file_path = os.path.join(download_dir, latest_file)
                    file_mtime = datetime.fromtimestamp(os.path.getmtime(download_file_path)).strftime('%Y-%m-%d %H:%M:%S')
                    print(f"从目录找到文件: {download_file_path}")
                    print(f"文件修改时间: {file_mtime}")
                    print(f"\n=== 准备发送邮件 ===")
                    log(f"下载完成(目录兜底)，开始编辑并发送邮件: {download_file_path}")
                    send_email_with_attachment(download_file_path)
                else:
                    print("未找到下载的Excel文件")
                    log("未找到下载的Excel文件")

        except Exception as e:
            print(f"\n出错: {e}")
            log(f"任务执行出错: {e}")
            import traceback
            traceback.print_exc()
        finally:
            browser.close()
            log("=== 任务执行结束 ===")


if __name__ == "__main__":
    main()
